# Excel Exporter - Excel 匯出功能
import logging
import os
import time
import json
import tempfile
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import Optional
from backend.utils.parser import extract_structured_data
from backend.repositories.job_repository import JobRepository

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Handles Excel export functionality using native openpyxl."""
    
    def __init__(self, project_repo):
        self.project_repo = project_repo

    @staticmethod
    def _sanitize_windows_filename_fragment(value: str) -> str:
        sanitized = re.sub(r'[<>:"/\\|?*]', "_", value or "")
        return sanitized.strip()
    
    async def run_excel(self, project_id: str):
        """Export project data to Excel file."""
        try:
            path = await self.archive_to_excel(project_id)
            return {"status": "excel_exported", "path": path}
        except Exception as e:
            logger.error(f"Error exporting excel for {project_id}: {e}")
            raise e

    async def archive_to_excel(self, project_id: str, excel_name: Optional[str] = None) -> str:
        """
        Export project jobs data to Excel file with main table and details table.
        
        Args:
            project_id: Project identifier
            excel_name: Optional custom Excel filename
            
        Returns:
            Path to the created Excel file
        """
        root = self.project_repo._project_root(project_id)
        if not root.exists():
            raise FileNotFoundError("project root not found")

        # 從全域集中資料庫讀取 (透過 JobRepository)
        from backend.database.core import AsyncSessionLocal
        job_repo = JobRepository(project_id, session_factory=AsyncSessionLocal)
        base_jobs_list = await job_repo.list_jobs()
        
        if not base_jobs_list:
            raise FileNotFoundError("No jobs found for this project")
            
        jobs_list = []
        for b_job in base_jobs_list:
            job_data = await job_repo.get_job(b_job["job_id"])
            if job_data:
                jobs_list.append(job_data)
        
        # Define columns layout
        main_cols = [
            "狀態",
            "檔名",
            "來源檔案(位置)",
            "VLM處理時間",
            "總時間",
            "備註",
            "檔案日期",
            "供應商",
            "金額",
            "人工修正",
            "VLM結果本文",
        ]
        
        detail_cols = [
            "狀態", "檔名", "來源檔案(位置)", "專案", "發票號碼", "供應商", 
            "報帳名目", "品項名稱", "數量", "單價", "小計"
        ]

        main_rows = []
        detail_rows = []

        for row in jobs_list:
            job_id = row.get("job_id")
            image_path = row.get("image_path")
            filename = Path(image_path).name if image_path else None
            
            # Parse stats from JSON fields
            vlm_time = 0
            try:
                vlm_stats = row.get("vlm_stats")
                if vlm_stats:
                    vlm_stats_data = json.loads(vlm_stats) if isinstance(vlm_stats, str) else vlm_stats
                    vlm_time = vlm_stats_data.get("total_time_s", 0)
            except (json.JSONDecodeError, TypeError):
                pass
            
            total = None
            try:
                if row.get("created_at") and row.get("updated_at"):
                    total = float(row.get("updated_at")) - float(row.get("created_at"))
            except Exception:
                total = None

            raw_vlm = row.get("vlm_result_json")
            job_status = row.get("status")

            human_correction_text = None
            vlm_body_text = None

            display_result = {}
            if job_id:
                try:
                    display_result = await job_repo.get_display_result(job_id) or {}
                except Exception:
                    display_result = {}
            
            # Parse VLM JSON to extract multiple pieces of information
            parsed_vlm = {}
            if isinstance(display_result, dict) and display_result:
                parsed_vlm = display_result
            elif isinstance(raw_vlm, str) and raw_vlm.strip():
                try:
                    parsed_vlm = json.loads(raw_vlm)
                except json.JSONDecodeError:
                    parsed_vlm = {}

            # Generate text from flat structure
            vlm_body_text = self._generate_text_from_vlm_result(parsed_vlm)

            # extract_structured_data handles flat structure
            structured = extract_structured_data(display_result)
            if not structured:
                structured = extract_structured_data(raw_vlm)

            supplier = structured.get("supplier") or row.get("supplier", "")
            total_amount = structured.get("total_amount")
            if total_amount in (None, ""):
                total_amount = row.get("total_amount", "")
            file_date = (
                structured.get("date", "")
                or structured.get("invoice_date", "")
                or row.get("invoice_date", "")
                or ""
            )

            main_rows.append(
                {
                    "狀態": job_status,
                    "檔名": filename,
                    "來源檔案(位置)": image_path,
                    "VLM處理時間": vlm_time,
                    "總時間": total,
                    "備註": None,
                    "檔案日期": file_date,
                    "供應商": supplier,
                    "金額": total_amount,
                    "人工修正": human_correction_text,
                    "VLM結果本文": vlm_body_text,
                }
            )

            # 細項：取 structured["items"]
            items = structured.get("items") or []
            for it in items:
                detail_rows.append(
                    {
                        "狀態": job_status,
                        "檔名": filename,
                        "來源檔案(位置)": image_path,
                        "專案": project_id,
                        "發票號碼": structured.get("invoice_id", "") or structured.get("voucher_id", "") or row.get("voucher_id", ""),
                        "供應商": supplier,
                        "報帳名目": it.get("category", ""),
                        "品項名稱": it.get("description", ""),
                        "數量": it.get("quantity", ""),
                        "單價": it.get("price", ""),
                        "小計": it.get("total", ""),
                    }
                )

        if not excel_name:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            project = await self.project_repo.get_project(project_id)
            project_name = ""
            if isinstance(project, dict):
                project_name = str(project.get("name") or "")
                if not project_name:
                    project_name = str((project.get("metadata") or {}).get("name") or "")

            safe_project_id = self._sanitize_windows_filename_fragment(project_id) or "UNKNOWN"
            safe_project_name = self._sanitize_windows_filename_fragment(project_name) or "未命名"
            excel_name = f"{safe_project_id}「{safe_project_name}」_預結算表_{ts}.xlsx"
        out_path = root / excel_name

        # Create workbook and worksheets
        wb = openpyxl.Workbook()
        
        # Main Sheet
        ws_main = wb.active
        ws_main.title = "主表"
        ws_main.append(main_cols)
        for r in main_rows:
            ws_main.append([r.get(c) for c in main_cols])

        # Detail Sheet
        ws_detail = wb.create_sheet("細項表")
        ws_detail.append(detail_cols)
        for r in detail_rows:
            ws_detail.append([r.get(c) for c in detail_cols])

        # Style Definitions
        header_font = Font(name="Microsoft JhengHei", size=11, bold=True)
        cell_font = Font(name="Microsoft JhengHei", size=10)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_side = Side(style='thin', color='D0D0D0')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for ws in [ws_main, ws_detail]:
            # Format headers
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Format data cells
            for r_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    cell.font = cell_font
                    cell.border = thin_border
                    val = cell.value
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    val_len = 0
                    for char in val:
                        val_len += 2 if ord(char) > 127 else 1
                    max_len = max(max_len, val_len)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

        # Write safely
        fd, tmp = tempfile.mkstemp(dir=str(out_path.parent), suffix=".xlsx")
        os.close(fd)
        try:
            wb.save(tmp)
            os.replace(tmp, str(out_path))
        except Exception as e:
            try:
                os.unlink(tmp)
            except Exception:
                pass
            raise

        # 更新全域狀態
        try:
            await self.project_repo.update_project_status(project_id, "ARCHIVED")
        except Exception:
            pass

        logger.info("archive_to_excel completed: %s", str(out_path))
        return str(out_path)
    
    def _generate_text_from_vlm_result(self, parsed_vlm: dict) -> str:
        """從扁平 VLM 結果生成文字摘要"""
        if not parsed_vlm:
            return ""
        
        lines = []
        
        header = parsed_vlm.get("header", {})
        if header.get("supplier"):
            lines.append(f"# {header['supplier']}")
        if header.get("invoice_id"):
            lines.append(f"發票號碼: {header['invoice_id']}")
        if header.get("date"):
            lines.append(f"日期: {header['date']}")
        
        items = parsed_vlm.get("items", [])
        if items:
            lines.append("")
            lines.append("| 名目 | 單價 | 數量 | 小計 | 品名 |")
            lines.append("|------|------|------|------|------|")
            for item in items:
                cat = item.get("category", "")
                price = item.get("price", "")
                qty = item.get("qty", item.get("quantity", ""))
                total = item.get("total", "")
                name = item.get("name", item.get("description", ""))
                lines.append(f"| {cat} | {price} | {qty} | {total} | {name} |")
        
        summary = parsed_vlm.get("summary", {})
        if summary.get("total"):
            lines.append("")
            lines.append(f"**合計**: {summary['total']}")
        
        return "\n".join(lines)

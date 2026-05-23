# Regeneration Handler - 人工修正重新生成功能
import logging
import time
import json
import openpyxl
from typing import Optional, Dict, Any

logger = logging.getLogger(__name__)


class RegenerationHandler:
    """Handles regeneration of LLM results from human corrections using openpyxl."""
    
    def __init__(self, project_repo, excel_exporter):
        self.project_repo = project_repo
        self.excel_exporter = excel_exporter
    
    async def regenerate_from_archive(
        self, project_id: str, excel_path: str, config: Dict[str, Any]
    ) -> Optional[str]:
        """
        Reads an archived Excel file, uses the '人工修正' column to regenerate LLM output,
        and creates a new archive. The status is updated to 'human_correct'.
        
        Args:
            project_id: Project identifier
            excel_path: Path to the archived Excel file
            config: LLM handler configuration
            
        Returns:
            Path to the new archive, or None on failure
        """
        from backend.processing.llm_handler import LLMHandler
        logger.info(f"Starting regeneration for project '{project_id}' from '{excel_path}'")

        root = self.project_repo._project_root(project_id)

        # --- Read the Excel file ---
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            if "主表" not in wb.sheetnames:
                logger.error(f"Sheet '主表' not found in Excel file '{excel_path}'")
                wb.close()
                return None
            ws = wb["主表"]
        except Exception as e:
            logger.error(f"Failed to read Excel file '{excel_path}': {e}")
            return None

        # Parse headers and rows
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            logger.error("Excel sheet '主表' is empty")
            wb.close()
            return None

        if not headers or "人工修正" not in headers:
            logger.error(f"Column '人工修正' not found in Excel file '{excel_path}'")
            wb.close()
            return None

        manual_idx = headers.index("人工修正")
        filename_idx = headers.index("檔名") if "檔名" in headers else -1

        rows_data = []
        for r in rows_iter:
            if not r:
                continue
            filename = r[filename_idx] if (filename_idx != -1 and filename_idx < len(r)) else None
            manual_correction = r[manual_idx] if (manual_idx < len(r)) else None

            # Skip empty entries
            if manual_correction is None or str(manual_correction).strip() == "" or filename is None or str(filename).strip() == "":
                continue

            rows_data.append({
                "檔名": str(filename).strip(),
                "人工修正": str(manual_correction).strip()
            })
        wb.close()

        # --- Initialize LLM Handler ---
        try:
            llm_handler = LLMHandler(config=config)
        except Exception as e:
            logger.error(f"Failed to initialize LLMHandler: {e}")
            return None
        
        # --- Process each row (使用全域集中 JobRepository) ---
        from backend.database.core import AsyncSessionLocal
        from backend.repositories.job_repository import JobRepository
        job_repo = JobRepository(project_id, session_factory=AsyncSessionLocal)
        
        for item in rows_data:
            filename = item["檔名"]
            manual_correction = item["人工修正"]

            # Find the job from the global DB
            all_jobs = await job_repo.list_jobs()
            matched_job = None
            for j in all_jobs:
                if j.get("image_path", "").endswith(filename):
                    matched_job = j
                    break
            
            if not matched_job:
                logger.warning(f"No job found in database for filename '{filename}'. Skipping.")
                continue
            
            job_id = matched_job["job_id"]
            logger.info(f"Processing job_id '{job_id}' for file '{filename}'...")

            # Regenerate structured data from LLM
            structured_part = llm_handler.regenerate_from_corrected_text(manual_correction)
            
            # Construct the final JSON object in the new flat format
            final_json_obj = {
                "receipt_type": structured_part.get("receipt_type", ""),
                "header": structured_part.get("header", {}),
                "items": structured_part.get("items", []),
                "summary": structured_part.get("summary", {}),
                "audit": {
                    "confidence": 1.0,
                    "issues": [],
                    "corrections": [{
                        "source": "human",
                        "timestamp": int(time.time()),
                        "description": "人工修正"
                    }]
                }
            }
            final_json_str = json.dumps(final_json_obj, ensure_ascii=False)

            # Update via JobRepository
            await job_repo.update_job(job_id, vlm_result_json=final_json_str, status="human_correct")
        
        logger.info("Finished processing all rows from Excel file.")
        
        # --- Re-archive to a new Excel file ---
        new_excel_name = f"{project_id}_regenerated_{int(time.time())}.xlsx"
        logger.info(f"Re-archiving results to '{new_excel_name}'...")
        try:
            new_archive_path = await self.excel_exporter.archive_to_excel(project_id, excel_name=new_excel_name)
            logger.info(f"Successfully created regenerated archive: {new_archive_path}")
            return new_archive_path
        except Exception as e:
            logger.error(f"Failed to re-archive project: {e}")
            return None
